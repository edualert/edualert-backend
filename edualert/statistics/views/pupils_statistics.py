from django.db.models import Q
from django.db.models.functions import Lower
from django.http import Http404
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles.numbers import FORMAT_TEXT
from rest_framework import generics
from rest_framework.generics import get_object_or_404
from rest_framework.response import Response
from rest_framework.views import APIView

from edualert.academic_calendars.utils import get_current_academic_calendar, get_second_semester_end_events
from edualert.catalogs.models import StudentCatalogPerYear
from edualert.catalogs.utils import get_behavior_grade_limit
from edualert.common.permissions import IsAdministratorOrSchoolEmployee, IsTeacher, IsPrincipal
from edualert.common.search_and_filters import CommonOrderingFilter
from edualert.profiles.models import UserProfile
from edualert.statistics.models import StudentAtRiskCounts
from edualert.statistics.pagination import StatisticsPagination
from edualert.statistics.serializers import PupilStatisticsForORSSerializer, PupilStatisticsForSchoolEmployeeSerializer, \
    StudentsAveragesSerializer, StudentsAbsencesSerializer, StudentsBehaviorGradeSerializer, \
    SchoolStudentAtRiskSerializer, StudentAtRiskSerializer


class PupilsStatistics(generics.ListAPIView):
    permission_classes = (IsAdministratorOrSchoolEmployee,)
    pagination_class = StatisticsPagination
    ordering_fields = [
        'student_name', 'avg_sem1', 'avg_sem2', 'avg_final',
        'unfounded_abs_count_sem1', 'unfounded_abs_count_sem2', 'unfounded_abs_count_annual',
        'second_examinations_count', 'behavior_grade_sem1', 'behavior_grade_sem2', 'behavior_grade_annual'
    ]
    ordering = None
    ordering_extra_annotations = {'student_name': Lower('student__full_name')}
    filterset_fields = ['academic_year', ]
    filter_backends = [DjangoFilterBackend, CommonOrderingFilter]

    def get_serializer_context(self):
        context = super().get_serializer_context()
        profile = self.request.user.user_profile
        if profile.school_unit:
            academic_profile = profile.school_unit.academic_profile
            context.update({
                'behavior_grade_limit': get_behavior_grade_limit(academic_profile),
            })
        return context

    def get_serializer_class(self):
        profile = self.request.user.user_profile

        if profile.user_role == UserProfile.UserRoles.ADMINISTRATOR:
            return PupilStatisticsForORSSerializer
        return PupilStatisticsForSchoolEmployeeSerializer

    def get_search_filter(self, profile):
        search = self.request.query_params.get('search')
        search_filter = Q()
        if search:
            if profile.user_role == UserProfile.UserRoles.ADMINISTRATOR:
                search_filter = Q(student__labels__text__unaccent__icontains=search) | \
                                Q(study_class__school_unit__name__unaccent__icontains=search)
            else:
                search_filter = Q(student__labels__text__unaccent__icontains=search) | \
                                Q(student__full_name__unaccent__icontains=search)

        return search_filter

    def get_filters(self):
        filters = {}

        generic_academic_program = self.request.query_params.get('academic_program')
        if generic_academic_program:
            try:
                program_id = int(generic_academic_program)
                filters['study_class__academic_program__generic_academic_program_id'] = program_id
            except ValueError:
                pass

        study_class_grade = self.request.query_params.get('study_class_grade')
        if study_class_grade:
            filters['study_class__class_grade'] = study_class_grade

        return filters

    def get_queryset(self):
        profile = self.request.user.user_profile

        search_filter = self.get_search_filter(profile)
        filters = self.get_filters()

        if profile.user_role == UserProfile.UserRoles.ADMINISTRATOR:
            self.ordering = ['student__id']
            queryset = StudentCatalogPerYear.objects.select_related('student__school_unit__academic_profile', 'study_class') \
                .prefetch_related('student__labels') \
                .filter(search_filter, **filters) \
                .distinct()
        else:
            self.ordering = ['student_name']

            if profile.user_role == UserProfile.UserRoles.TEACHER:
                filters['study_class__teachers'] = profile

            queryset = StudentCatalogPerYear.objects.select_related('student__school_unit__academic_profile', 'study_class') \
                .prefetch_related('student__labels') \
                .filter(search_filter, study_class__school_unit_id=profile.school_unit_id, **filters) \
                .distinct()

        return queryset


class OwnStudentsStatisticsBaseClass(generics.ListAPIView):
    permission_classes = (IsTeacher,)
    pagination_class = StatisticsPagination

    def get_queryset(self):
        current_calendar = get_current_academic_calendar()
        if not current_calendar:
            raise Http404()

        mastering_study_class = get_object_or_404(
            self.request.user.user_profile.mastering_study_classes,
            academic_year=current_calendar.academic_year
        )

        order_by = self.get_order_by(current_calendar)

        return mastering_study_class.student_catalogs_per_year \
            .select_related('student') \
            .order_by(order_by, Lower('student__full_name'))


class OwnStudentsAverages(OwnStudentsStatisticsBaseClass):
    serializer_class = StudentsAveragesSerializer

    @staticmethod
    def get_order_by(current_calendar):
        return '-avg_sem1' if timezone.now().date() < current_calendar.second_semester.ends_at else '-avg_final'


class OwnStudentsAbsences(OwnStudentsStatisticsBaseClass):
    serializer_class = StudentsAbsencesSerializer

    @staticmethod
    def get_order_by(current_calendar):
        return '-unfounded_abs_count_sem1' if timezone.now().date() < current_calendar.second_semester.ends_at else '-unfounded_abs_count_annual'


class OwnStudentsBehaviorGrades(OwnStudentsStatisticsBaseClass):
    serializer_class = StudentsBehaviorGradeSerializer

    def get_serializer_context(self):
        context = super().get_serializer_context()
        academic_profile = self.request.user.user_profile.school_unit.academic_profile
        context.update({
            'behavior_grade_limit': get_behavior_grade_limit(academic_profile),
        })
        return context

    @staticmethod
    def get_order_by(current_calendar):
        return '-behavior_grade_sem1' if timezone.now().date() < current_calendar.second_semester.ends_at else '-behavior_grade_annual'


class SchoolStudentsAtRisk(generics.ListAPIView):
    permission_classes = (IsPrincipal,)
    pagination_class = StatisticsPagination
    serializer_class = SchoolStudentAtRiskSerializer

    def get_serializer_context(self):
        context = super().get_serializer_context()
        academic_profile = self.request.user.user_profile.school_unit.academic_profile
        context.update({
            'behavior_grade_limit': get_behavior_grade_limit(academic_profile),
        })
        return context

    def get_queryset(self):
        current_calendar = get_current_academic_calendar()
        if not current_calendar:
            return StudentCatalogPerYear.objects.none()

        order_by1 = '-unfounded_abs_count_sem1' if timezone.now().date() < current_calendar.second_semester.ends_at else '-unfounded_abs_count_annual'
        order_by2 = '-behavior_grade_sem1' if timezone.now().date() < current_calendar.second_semester.ends_at else '-behavior_grade_annual'

        profile = self.request.user.user_profile
        return StudentCatalogPerYear.objects.select_related('student', 'study_class') \
            .filter(student__school_unit_id=profile.school_unit_id,
                    academic_year=current_calendar.academic_year,
                    student__is_at_risk=True) \
            .distinct() \
            .order_by(order_by1, order_by2, Lower('student__full_name'))


class SchoolStudentsAtRiskExport(APIView):
    permission_classes = (IsPrincipal,)

    def get(self, request, *args, **kwargs):
        from edualert.catalogs.utils import has_technological_category
        from edualert.catalogs.utils import get_current_semester
        from edualert.catalogs.models import StudentCatalogPerSubject
        from django.http import HttpResponse

        now = timezone.now()
        file_name = f"RaportStudentiRisc_{now.year}-{now.month}-{now.day}_{now.hour}:{now.minute}:{now.second}.csv"
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'

        profile = self.request.user.user_profile
        current_calendar = get_current_academic_calendar()
        if not current_calendar:
            return response

        second_semester_end_events = get_second_semester_end_events(current_calendar)
        is_technological_school = has_technological_category(profile.school_unit)

        current_row = 1
        headers = ['Nume', 'Medie Matematică', 'Medie Limba Română', 'Absențe nemotivate', 'Notă purtare',
                   'Telefon elev', 'Telefon părinți', 'Clasă', 'Descriere risc']
        column_widths = []

        workbook = Workbook()
        worksheet = workbook.active

        # write headers
        for i, header in enumerate(headers):
            cell = worksheet.cell(column=i + 1, row=current_row)
            cell.value = header
            column_widths.append(len(header))
        current_row += 1

        student_catalogs = StudentCatalogPerYear.objects.select_related('student', 'study_class') \
            .filter(student__school_unit_id=profile.school_unit_id,
                    academic_year=current_calendar.academic_year,
                    student__is_at_risk=True) \
            .distinct() \
            .order_by(Lower('student__full_name'))

        for student_catalog in student_catalogs:  # type: StudentCatalogPerYear
            student = student_catalog.student  # type: UserProfile
            study_class = student_catalog.study_class

            current_semester = get_current_semester(now.date(), current_calendar, second_semester_end_events,
                                                    study_class.class_grade_arabic, is_technological_school)

            math = StudentCatalogPerSubject.objects.filter(
                student_id=student.id, academic_year=current_calendar.academic_year, subject_name='Matematică').first()
            if not math:
                math = type('', (object,), dict(avg_annual='-', avg_sem1='-'))

            romanian = StudentCatalogPerSubject.objects.filter(
                Q(subject_name='Limba Română') | Q(subject_name='Limba și literatura română'), student_id=student.id,
                academic_year=current_calendar.academic_year).first()
            if not romanian:
                romanian = type('', (object,), dict(avg_annual='-', avg_sem1='-'))

            parents_phone_numbers = student.parents.values_list('phone_number', flat=True)
            filtered_parents_phone_numbers = filter(lambda x: x and x != '', parents_phone_numbers)

            row = [
                student.full_name,
                math.avg_annual,
                romanian.avg_annual,
                student_catalog.unfounded_abs_count_annual,
                student_catalog.behavior_grade_annual,
                student.phone_number,
                ','.join(filtered_parents_phone_numbers),
                study_class.class_grade + ' ' + study_class.class_letter,
                student.risk_description,
            ]
            if current_semester == 1:
                row[1] = '-'
                row[2] = '-'
                row[3] = student_catalog.unfounded_abs_count_sem1
                row[4] = '-'
            elif current_semester == 2:
                row[1] = math.avg_sem1
                row[2] = romanian.avg_sem1
                row[3] = student_catalog.unfounded_abs_count_sem2
                row[4] = student_catalog.behavior_grade_sem1

            # set the value and styling for each cell
            for i, val in enumerate(row):
                cell = worksheet.cell(column=i + 1, row=current_row)
                cell.alignment = Alignment(horizontal='left')
                cell.number_format = FORMAT_TEXT
                if val and val != '':
                    cell.value = val
                else:
                    cell.value = '-'

                # update column width if necessary
                width = len(str(cell.value))
                if width > column_widths[i]:
                    column_widths[i] = width
            current_row += 1

        # update the width of the columns
        for i, header in enumerate(headers):
            col_letter = chr(ord('A') + i)
            worksheet.column_dimensions[col_letter].width = column_widths[i]

        workbook.save(response)
        return response


class OwnStudentsAtRisk(generics.ListAPIView):
    permission_classes = (IsTeacher,)
    pagination_class = StatisticsPagination
    serializer_class = StudentAtRiskSerializer

    def get_serializer_context(self):
        context = super().get_serializer_context()
        academic_profile = self.request.user.user_profile.school_unit.academic_profile
        context.update({
            'behavior_grade_limit': get_behavior_grade_limit(academic_profile),
        })
        return context

    def get_queryset(self):
        current_calendar = get_current_academic_calendar()
        if not current_calendar:
            raise Http404()

        mastering_study_class = get_object_or_404(
            self.request.user.user_profile.mastering_study_classes,
            academic_year=current_calendar.academic_year
        )

        order_by1 = '-unfounded_abs_count_sem1' if timezone.now().date() < current_calendar.second_semester.ends_at else '-unfounded_abs_count_annual'
        order_by2 = '-behavior_grade_sem1' if timezone.now().date() < current_calendar.second_semester.ends_at else '-behavior_grade_annual'

        return mastering_study_class.student_catalogs_per_year.select_related('student') \
            .filter(student__is_at_risk=True) \
            .order_by(order_by1, order_by2, Lower('student__full_name'))


class StudentsRiskEvolution(APIView):
    permission_classes = (IsAdministratorOrSchoolEmployee,)

    def get(self, request, *args, **kwargs):
        today = timezone.now().date()
        month = self.request.GET.get('month')
        filters = {}

        if month and not (month.isnumeric() and 1 <= int(month) <= 12) or not month:
            month = today.month

        profile = self.request.user.user_profile
        if profile.user_role == UserProfile.UserRoles.ADMINISTRATOR:
            school_unit = self.request.GET.get('school_unit')
            filters.update({'school_unit_id': school_unit} if school_unit and school_unit.isnumeric() else {'by_country': True})

        elif profile.user_role == UserProfile.UserRoles.PRINCIPAL:
            filters.update({'school_unit_id': profile.school_unit_id})

        elif profile.user_role == UserProfile.UserRoles.TEACHER:
            current_calendar = get_current_academic_calendar()
            if current_calendar:
                filters.update({'study_class__class_master_id': profile.id, 'study_class__academic_year': current_calendar.academic_year})

        stats = StudentAtRiskCounts.objects.filter(year=today.year, month=month, **filters).first()
        return Response(stats.daily_counts if stats else [])
