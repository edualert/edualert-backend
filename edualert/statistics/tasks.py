import datetime
import math
import tempfile
from calendar import monthrange
from os import unlink

from celery import shared_task
from dateutil.relativedelta import relativedelta
from django.conf import settings
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Count, Q
from django.template.loader import get_template
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.styles.numbers import FORMAT_TEXT

from edualert.academic_calendars.utils import get_current_academic_calendar
from edualert.catalogs.models import StudentCatalogPerSubject
from edualert.common.constants import WEEKDAYS_MAP
from edualert.notifications.utils.emails import send_mail_with_attachments
from edualert.schools.models import RegisteredSchoolUnit
from edualert.statistics.models import StudentAtRiskCounts, SchoolUnitEnrollmentStats
from edualert.study_classes.models import StudyClass


@shared_task
def create_students_at_risk_counts_task():
    today = timezone.now().date()
    days_in_month = monthrange(today.year, today.month)[1]
    current_calendar = get_current_academic_calendar()
    if not current_calendar:
        return

    for school_unit in RegisteredSchoolUnit.objects.all():
        StudentAtRiskCounts.objects.create(
            month=today.month,
            year=today.year,
            school_unit=school_unit,
            daily_counts=[
                {
                    'day': day,
                    'weekday': WEEKDAYS_MAP[datetime.datetime(today.year, today.month, day).weekday()],
                    'count': 0
                } for day in range(1, days_in_month + 1)
            ]
        )
        for study_class in school_unit.study_classes.filter(academic_year=current_calendar.academic_year):
            StudentAtRiskCounts.objects.create(
                month=today.month,
                year=today.year,
                study_class=study_class,
                daily_counts=[
                    {
                        'day': day,
                        'weekday': WEEKDAYS_MAP[datetime.datetime(today.year, today.month, day).weekday()],
                        'count': 0
                    } for day in range(1, days_in_month + 1)
                ]
            )

    StudentAtRiskCounts.objects.create(
        month=today.month,
        year=today.year,
        by_country=True,
        daily_counts=[
            {
                'day': day,
                'weekday': WEEKDAYS_MAP[datetime.datetime(today.year, today.month, day).weekday()],
                'count': 0
            } for day in range(1, days_in_month + 1)
        ]
    )


@shared_task
def create_students_at_risk_counts_for_school_unit_task(school_unit_id):
    try:
        school_unit = RegisteredSchoolUnit.objects.get(id=school_unit_id)
    except ObjectDoesNotExist:
        return

    current_calendar = get_current_academic_calendar()
    if not current_calendar:
        return

    months_since_academic_year_start = get_months_since_academic_calendar_start(current_calendar)
    objects_to_create = []
    for year, month in months_since_academic_year_start:
        days_in_month = monthrange(year, month)[1]
        if not StudentAtRiskCounts.objects.filter(month=month, year=year, school_unit=school_unit).exists():
            objects_to_create.append(
                StudentAtRiskCounts(
                    month=month,
                    year=year,
                    school_unit=school_unit,
                    daily_counts=[
                        {
                            'count': 0,
                            'day': day,
                            'weekday': WEEKDAYS_MAP[datetime.datetime(year, month, day).weekday()]
                        } for day in range(1, days_in_month + 1)
                    ]
                )
            )
    StudentAtRiskCounts.objects.bulk_create(objects_to_create)


@shared_task
def create_students_at_risk_counts_for_study_class_task(study_class_id):
    try:
        study_class = StudyClass.objects.get(id=study_class_id)
    except ObjectDoesNotExist:
        return

    current_calendar = get_current_academic_calendar()
    if not current_calendar:
        return

    months_since_academic_year_start = get_months_since_academic_calendar_start(current_calendar)
    objects_to_create = []
    for year, month in months_since_academic_year_start:
        days_in_month = monthrange(year, month)[1]
        if not StudentAtRiskCounts.objects.filter(month=month, year=year, study_class=study_class).exists():
            objects_to_create.append(
                StudentAtRiskCounts(
                    month=month,
                    year=year,
                    study_class=study_class,
                    daily_counts=[
                        {
                            'count': 0,
                            'day': day,
                            'weekday': WEEKDAYS_MAP[datetime.datetime(year, month, day).weekday()]
                        } for day in range(1, days_in_month + 1)
                    ]
                )
            )
    StudentAtRiskCounts.objects.bulk_create(objects_to_create)


@shared_task
def create_school_unit_enrollment_stats_task():
    today = timezone.now().date()
    days_in_month = monthrange(today.year, today.month)[1]
    if not SchoolUnitEnrollmentStats.objects.filter(month=today.month, year=today.year).exists():
        SchoolUnitEnrollmentStats.objects.create(
            month=today.month,
            year=today.year,
            daily_statistics=[{
                'count': 0,
                'day': day,
                'weekday': WEEKDAYS_MAP[datetime.datetime(today.year, today.month, day).weekday()]
            } for day in range(1, days_in_month + 1)]
        )


@shared_task
def update_school_unit_enrollment_stats_task():
    today = timezone.now().date()

    stats = SchoolUnitEnrollmentStats.objects.get_or_create(month=today.month, year=today.year)[0]

    for index, daily_stat in enumerate(stats.daily_statistics):
        if daily_stat['day'] == today.day:
            stats.daily_statistics[index]['count'] += 1
            stats.save()
            break


def get_months_since_academic_calendar_start(current_calendar):
    today = timezone.now().date()
    date = current_calendar.first_semester.starts_at.replace(day=1)

    months = []
    while date <= today:
        months.append((date.year, date.month))
        date += relativedelta(months=1)

    return months


def _get_month_name(month_number):
    month_names = ['Ianuarie', 'Februarie', 'Martie', 'Aprilie', 'Mai', 'Iunie', 'Iulie', 'August', 'Septembrie',
                   'Octombrie', 'Noiembrie', 'Decembrie']
    # the list is index from 0 and the `month` variable from 1
    return month_names[month_number - 1]


@shared_task
def send_monthly_school_unit_absence_report_task():
    today = timezone.now().date()
    reported_date = today - relativedelta(months=1)
    academic_calendar = get_current_academic_calendar()
    month_name = _get_month_name(reported_date.month)

    if not academic_calendar:
        return

    # generate all the reports which have to be sent
    report_files = []
    try:
        _generate_report_files(report_files, academic_calendar.academic_year, reported_date, today)
    except Exception as err:
        # remove all the generated temporary files before bubbling up the exception
        for _, file in report_files:
            unlink(file.name)
        raise err

    # Deliver all the reports to all the assigned emails
    delivery_emails = settings.ABSENCES_REPORT_DELIVERY_EMAILS
    try:
        subject = 'Raport lunar absențe - {} {}'.format(month_name, reported_date.year)
        content = 'Bună ziua!\n\nAcesta este un raport lunar automat în care veți găsi atașate documentele ' \
                  'care conțin evidența absențelor pentru {} {}.'.format(month_name, reported_date.year)
        # compose attachments based on: filesystem filename, attachment filename
        attachments = [(file.name, '{}.xlsx'.format(school_name)) for school_name, file in report_files]

        template = get_template('message.html')
        template_context = {
            'title': subject, 'body': content, 'show_my_account': False, 'signature': 'Echipa EduAlert'
        }
        bodies = {
            'text/html': template.render(context=template_context)
        }
        send_mail_with_attachments(subject, bodies, settings.SERVER_EMAIL, delivery_emails, attachments)
    finally:
        # remove all the generated temporary files
        for _, file in report_files:
            unlink(file.name)


def _generate_report_files(report_files, academic_year, reported_date, current_date):
    query_offset = 0
    query_batch_size = 200
    school_units_count = RegisteredSchoolUnit.objects.count()
    school_units_qs = RegisteredSchoolUnit.objects.all().order_by('created')

    # Iterate over all registered school units in batches of `query_batch_size` to avoid running out of memory in
    # case we have a lot of registered school units.
    #
    # Warning!
    # This comes with all the drawbacks of pagination. Meaning there exists the possibility of iterating over an
    # item twice or not at all when a new item is inserter before our current index or an item is removed before
    # our current index.
    while query_offset < school_units_count:
        for school_unit in school_units_qs[query_offset:query_offset + query_batch_size]:
            # generate a temporary file for this school unit and add it to the list
            file = tempfile.NamedTemporaryFile(prefix='edu_report_', suffix='.xslx', delete=False)
            report_files.append((school_unit.name, file))
            # compute report data and write it to the temporary file
            report_data = _compute_report_data_for_school_unit(academic_year, reported_date, school_unit)
            _write_report_xslx(file.name, reported_date, current_date, report_data.values())
            # move on to the next bach
            query_offset += query_batch_size


def _compute_report_data_for_school_unit(academic_year, reported_date, school_unit):
    year = reported_date.year
    month = reported_date.month
    unfounded_absences = Count('absence', filter=Q(absence__is_founded=False, absence__taken_at__year=year, absence__taken_at__month=month))
    founded_absences = Count('absence', filter=Q(absence__is_founded=True, absence__taken_at__year=year, absence__taken_at__month=month))

    classes_query = StudentCatalogPerSubject.objects \
        .filter(academic_year=academic_year, study_class__school_unit_id=school_unit.id) \
        .values('subject_name', 'study_class__class_grade', 'study_class__class_grade_arabic',
                'study_class__class_letter', 'study_class__class_master__full_name') \
        .annotate(unfounded_absences=unfounded_absences, founded_absences=founded_absences) \
        .order_by('study_class__class_grade_arabic', 'study_class__class_letter', 'subject_name')

    # transform query results
    classes = {}
    for row in classes_query:
        class_name = '{} {}'.format(row['study_class__class_grade'], row['study_class__class_letter'])

        # add class if it doesn't exist
        if not classes.get(class_name):
            classes[class_name] = {
                'class_name': class_name,
                'class_master': row['study_class__class_master__full_name'],
            }

        # add subjects if the don't exist
        clazz = classes[class_name]
        if not clazz.get('subjects'):
            clazz['subjects'] = []

        # add the subject to the list
        clazz['subjects'].append({
            'subject_name': row['subject_name'],
            'founded_absences': row['founded_absences'],
            'unfounded_absences': row['unfounded_absences'],
        })

    return classes


def _write_report_xslx(filename, reported_date, current_date, classes):
    def set_border(cell):
        cell.border = Border(
            left=Side(border_style='thin', color='FF000000'),
            right=Side(border_style='thin', color='FF000000'),
            top=Side(border_style='thin', color='FF000000'),
            bottom=Side(border_style='thin', color='FF000000'),
        )

    workbook = Workbook()
    first_sheet = True

    for clazz in classes:
        class_name = clazz['class_name']
        class_master = clazz['class_master']

        if first_sheet:
            # change the title of the first sheet
            worksheet = workbook.active
            worksheet.title = class_name
            first_sheet = False
        else:
            worksheet = workbook.create_sheet(class_name)

        # add title
        worksheet['A1'] = 'FISA EVIDENTA ABSENTE'
        worksheet['A1'].font = Font(name='Calibri', bold=True)
        worksheet['A1'].alignment = Alignment(horizontal='center')
        worksheet.merge_cells('A1:F1')

        # add subtitle
        worksheet['A2'] = 'Luna {} {}'.format(_get_month_name(reported_date.month), reported_date.year)
        worksheet['A2'].font = Font(name='Calibri', bold=True)
        worksheet['A2'].alignment = Alignment(horizontal='center')
        worksheet.merge_cells('A2:F2')

        # add deadline
        worksheet['A3'] = 'Termen de predare: 15.{}.{}'.format(current_date.month, current_date.year)
        worksheet['A3'].font = Font(name='Calibri', bold=True)
        worksheet.merge_cells('A3:C3')

        worksheet['A4'] = '=CONCATENATE("Data predării: ", Cap!A1)'
        worksheet['A4'].font = Font(name='Calibri', bold=True)
        worksheet.merge_cells('A4:C4')

        # add study class
        worksheet['D3'] = 'Clasa:'
        worksheet['D3'].alignment = Alignment(horizontal='right')
        worksheet['E3'] = class_name

        # add class master
        worksheet['D4'] = 'Diriginte:'
        worksheet['D4'].font = Font(name='Calibri', bold=True)
        worksheet['D4'].alignment = Alignment(horizontal='right')

        worksheet['E4'] = class_master
        worksheet.merge_cells('E4:F4')

        # add statistics fields
        worksheet['A6'] = 'Numar elevi existenti la inceputul lunii:'
        worksheet.merge_cells('A6:B6')
        worksheet['A7'] = 'Numar elevi exmatriculati:'
        worksheet.merge_cells('A7:B7')
        worksheet['A8'] = 'Numar elevi retrasi:'
        worksheet.merge_cells('A8:B8')
        worksheet['A9'] = 'Numar elevi veniti prin transfer:'
        worksheet.merge_cells('A9:B9')
        worksheet['A10'] = 'Numar elevi plecati prin transfer:'
        worksheet.merge_cells('A10:B10')

        for cell in ['C6', 'C7', 'C8', 'C9', 'C10']:
            set_border(worksheet[cell])

        # write table headers
        worksheet['A12'] = 'Nr. crt'
        worksheet['B12'] = 'Disciplina'
        worksheet.merge_cells('B12:C12')
        worksheet['D12'] = 'Nr. abs. motivate'
        worksheet['E12'] = 'Nr. abs. nemotivate'
        worksheet['F12'] = 'Total absente'

        for cell in ['A12', 'B12', 'D12', 'E12', 'F12']:
            worksheet[cell].font = Font(name='Calibri', bold=True)
            worksheet[cell].alignment = Alignment(horizontal='center')
            set_border(worksheet[cell])

        # setup column widths
        worksheet.column_dimensions['A'].width = 7
        worksheet.column_dimensions['C'].width = 4.5

        max_width = max([len(worksheet[cell].value) for cell in ['A6', 'A7', 'A8', 'A9', 'A10']])
        worksheet.column_dimensions['B'].width = max_width - worksheet.column_dimensions['A'].width

        worksheet.column_dimensions['D'].width = len(worksheet['D12'].value)
        worksheet.column_dimensions['E'].width = len(worksheet['E12'].value)
        worksheet.column_dimensions['F'].width = len(worksheet['F12'].value)

        # begin writing table data
        total_founded_absences = 0
        total_unfounded_absences = 0
        row = 13
        study_width = worksheet.column_dimensions['B'].width + worksheet.column_dimensions['C'].width
        for index, study in enumerate(clazz['subjects']):
            study_name = study['subject_name']
            founded_absences = study['founded_absences']
            unfounded_absences = study['unfounded_absences']

            # sum absences for Total row
            total_founded_absences += founded_absences
            total_unfounded_absences += unfounded_absences

            # write subject row
            worksheet.cell(column=1, row=row, value=index + 1)
            cell = worksheet.cell(column=2, row=row, value=study_name)
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            worksheet.merge_cells('B{}:C{}'.format(row, row))
            worksheet.cell(column=4, row=row, value=founded_absences)
            worksheet.cell(column=5, row=row, value=unfounded_absences)
            worksheet.cell(column=6, row=row, value=founded_absences + unfounded_absences)

            # adjust height if necessary
            if len(str(cell.value)) > study_width:
                height = math.ceil(len(cell.value) / study_width) * 15
                worksheet.row_dimensions[row].height = height

            # set the borders for the cells
            for column in range(1, 7):
                cell = worksheet.cell(column=column, row=row)
                set_border(cell)
            # move to next row
            row += 1

        # write Total row
        cell = worksheet.cell(column=1, row=row, value='TOTAL')
        cell.font = Font(name='Calibri', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        worksheet.cell(column=4, row=row, value=total_founded_absences)
        worksheet.cell(column=5, row=row, value=total_unfounded_absences)
        worksheet.cell(column=6, row=row, value=total_founded_absences + total_unfounded_absences)
        # set the borders for the cells
        for column in range(1, 7):
            cell = worksheet.cell(column=column, row=row)
            set_border(cell)

    # add sheet to allow easier change for global values
    worksheet = workbook.create_sheet('Cap')
    worksheet['A1'] = 'Setat din Cap:A1'
    worksheet['A1'].number_format = FORMAT_TEXT

    worksheet.column_dimensions['A'].width = len(str(worksheet['A1'].value))

    workbook.save(filename)
